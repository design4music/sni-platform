"""Export all unique signal values by category to Excel."""

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

COLUMNS = [
    "persons",
    "orgs",
    "places",
    "commodities",
    "policies",
    "systems",
    "named_events",
]

conn = psycopg2.connect("postgresql://postgres@localhost:5432/sni_v2")
cur = conn.cursor()

wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for col in COLUMNS:
    cur.execute(
        f"""
        SELECT val, COUNT(*)::int as title_count
        FROM title_labels, unnest({col}) AS val
        WHERE {col} IS NOT NULL
        GROUP BY val
        ORDER BY title_count DESC
    """
    )
    rows = cur.fetchall()

    ws = wb.create_sheet(title=col)
    ws.append(["Signal Value", "Title Count"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14

    for val, count in rows:
        ws.append([val, count])

    print(f"{col}: {len(rows)} unique values")

cur.close()
conn.close()

out = "out/signal_values.xlsx"
wb.save(out)
print(f"\nSaved to {out}")
